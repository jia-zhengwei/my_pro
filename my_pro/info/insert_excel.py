# coding: utf-8

from .database_relation import *
from .mysql_con import OperateDatabase
from openpyxl.workbook import Workbook

operation_databases = OperateDatabase()


def output_process():
    """
    从数据库获取数据插入到（过程流程）excel表格中
    """

    wb = Workbook()
    ws = wb.active
    start_row = 1
    start_col = 1
    with operation_databases.session_scope() as session:
        processes = session.query(Process).all()
        for process in processes:
            end_row = start_row
            end_col = start_col
            ws.cell(row=end_row, column=end_col).value = process.process_code
            ws.cell(row=end_row, column=end_col + 1).value = process.process_name
            process_flow = session.query(ProcessFlow).filter(ProcessFlow.id == process.process_flow_id).first()
            try:
                if process_flow.procezo:
                    ws.cell(row=end_row, column=end_col + 2).value = process_flow.procezo
                    if process_flow.move:
                        ws.cell(row=end_row, column=end_col + 3).value = process_flow.move
                    if process_flow.storage:
                        ws.cell(row=end_row, column=end_col + 4).value = process_flow.storage
                    if process_flow.inspect:
                        ws.cell(row=end_row, column=end_col + 5).value = process_flow.inspect
            except AttributeError:
                continue
            product_features = session.query(ProductFeature).filter(ProductFeature.process_id == process.id).all()
            for product_feature in product_features:
                ws.cell(row=end_row, column=end_col + 6).value = product_feature.product_features
                end_row += 1
            end_row = start_row
            procedure_features = session.query(ProcedureVariation).filter(
                ProcedureVariation.process_id == process.id).all()
            for procedure_feature in procedure_features:
                print(end_row)
                ws.cell(row=end_row, column=end_col + 7).value = procedure_feature.procedure_variation
                end_row += 1
            start_row = start_row + len(procedure_features) + 1
        wb.save('过程流程.xlsx')


def output_pfmea():
    """
    从数据库获取数据插入到pfmea的excel文件
    """

    wb = Workbook()
    ws = wb.active
    start_row = 1
    start_col = 1
    with operation_databases.session_scope() as session:
        processes = session.query(Process).all()
        for process in processes:
            end_row = start_row
            end_col = start_col
            ws.cell(row=end_row, column=end_col).value = process.process_code
            requirements = session.query(Requirement).filter(Requirement.process_id == process.id).all()
            for requirement in requirements:
                requirement_row = start_row
                ws.cell(row=requirement_row, column=end_col + 1).value = requirement.requirement
                failure_modes = session.query(FailureMode).filter(FailureMode.requirement_id == requirement.id).all()
                for failure_mode in failure_modes:
                    mode_row = start_row
                    ws.cell(row=mode_row, column=end_col + 2).value = failure_mode.failure_mode_name
                    ws.cell(row=mode_row, column=end_col + 5).value = failure_mode.classification
                    failure_effects = session.query(FailEffects).filter(
                        FailEffects.failure_mode_id == failure_mode.id).all()
                    failure_causes = session.query(FailureCauses).filter(
                        FailureCauses.failure_mode == failure_mode.id).all()
                    causes_row = start_row
                    effect_row = start_row
                    if len(failure_causes) >= len(failure_effects):
                        for failure_effect in failure_effects:
                            ws.cell(row=effect_row, column=end_col + 3).value = failure_effect.effects_content
                            ws.cell(row=effect_row, column=end_col + 4).value = failure_effect.severity
                            effect_row += 1
                        for failure_cause in failure_causes:
                            ws.cell(row=causes_row, column=end_col + 6).value = failure_cause.causes_content
                            ws.cell(row=causes_row, column=end_col + 7).value = failure_cause.control_prevention
                            ws.cell(row=causes_row, column=end_col + 8).value = failure_cause.occurrence
                            ws.cell(row=causes_row, column=end_col + 9).value = failure_cause.control_detection
                            ws.cell(row=causes_row, column=end_col + 10).value = failure_cause.detection
                            ws.cell(row=causes_row, column=end_col + 11).value = failure_cause.RPN
                            causes_row += 1
                        start_row = start_row + len(failure_causes) + 1
        wb.save('pfmea.xlsx')


def insert_control():
    """
    从数据库中获取数据插入到（控制清单）excel
    """
    wb = Workbook()
    ws = wb.active
    start_row = 1
    start_col = 1
    with operation_databases.session_scope() as session:
        processes = session.query(Process).all()
        for process in processes:
            end_row = start_row
            end_col = start_col
            ws.cell(row=end_row, column=end_col).value = process.process_code
            ws.cell(row=end_row, column=end_col + 1).value = process.process_name
            machines = session.query(Machine).filter(Machine.process_id == process.id).all()
            products = session.query(ProductFeature).filter(ProductFeature.process_id == process.id
                                                            and ProductFeature.excel_control == 1).all()
            for machine in machines:
                machine_row = start_row
                ws.cell(row=machine_row, column=end_col + 2).value = machine.machine_name
                machine_row += 1
            for product in products:
                pro_row = start_row
                ws.cell(row=pro_row, column=end_col + 4).value = product.product_features
                pass


if __name__ == '__main__':
    output_pfmea()
